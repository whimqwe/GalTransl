from typing import Dict, List
import openpyxl
from os.path import join as joinpath, splitext
from GalTransl.CSplitter import SplitChunkMetadata
from GalTransl import LOGGER
from GalTransl.ConfigHelper import CProjectConfig
import csv
from InquirerPy import inquirer
from InquirerPy.base.control import Choice
import asyncio
import sys  # Added for input
import os


def load_name_table(
    name_table_path: str, firstime_load: bool, chunks: List[SplitChunkMetadata], proj_config: CProjectConfig
) -> Dict[str, str]:
    """
    This function loads the name table from the given path, supporting both .xlsx and .csv formats.
    It prompts the user to edit the file and reloads ONCE if CN_Name entries are missing.

    Args:
    - name_table_path: The path to the name table (.xlsx or .csv).

    Returns:
    - A dictionary containing the name table.
    """

    usePostDictInName = proj_config.getDictCfgSection("usePostDictInName") or False
    useGPTDictInName = proj_config.getDictCfgSection("useGPTDictInName") or False
    usePreDictInName=proj_config.getDictCfgSection("usePreDictInName") or False
    gpt_dic = proj_config.gpt_dic
    post_dic = proj_config.post_dic
    pre_dic=proj_config.pre_dic

    def _load_internal(path: str) -> tuple[Dict[str, str], List[str], bool]:
        """Internal helper to load and check the name table."""
        name_table_internal: Dict[str, str] = {}
        missing_cn_names_internal: List[str] = []
        _, file_extension = splitext(path)
        file_extension = file_extension.lower()
        file_loaded_successfully_internal = False

        try:
            if file_extension == ".xlsx":
                workbook = openpyxl.load_workbook(path)
                sheet = workbook.active
                header = [cell.value for cell in sheet[1]]
                try:
                    jp_name_col_idx = header.index("JP_Name")
                    cn_name_col_idx = header.index("CN_Name")
                except ValueError:
                    LOGGER.warning(f"name替换表 {path} 缺少 'JP_Name' 或 'CN_Name' 列")
                    return {}, [], False

                for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
                    jp_name_cell = row[jp_name_col_idx]
                    cn_name_cell = row[cn_name_col_idx]
                    jp_name = jp_name_cell.value
                    cn_name = cn_name_cell.value

                    if jp_name is not None:
                        jp_name_str = str(jp_name)
                        if cn_name is None or str(cn_name).strip() == "":
                            missing_cn_names_internal.append(jp_name_str)
                        else:
                            name_table_internal[jp_name_str] = str(cn_name)
                file_loaded_successfully_internal = True

            elif file_extension == ".csv":
                with open(path, "r", newline="", encoding="utf-8-sig") as csvfile:
                    reader = csv.reader(csvfile)
                    try:
                        header = next(reader)
                    except StopIteration:
                        LOGGER.warning(f"CSV name替换表 {path} 为空或无法读取表头")
                        return {}, [], False
                    try:
                        jp_name_col_idx = header.index("JP_Name")
                        cn_name_col_idx = header.index("CN_Name")
                    except ValueError:
                        LOGGER.warning(
                            f"CSV name替换表 {path} 缺少 'JP_Name' 或 'CN_Name' 列"
                        )
                        return {}, [], False

                    for row_idx, row in enumerate(reader, start=2):
                        if len(row) > max(jp_name_col_idx, cn_name_col_idx):
                            jp_name = row[jp_name_col_idx]
                            cn_name = row[cn_name_col_idx]

                            if jp_name is not None:
                                jp_name_str = str(jp_name)
                                if cn_name is None or str(cn_name).strip() == "":
                                    missing_cn_names_internal.append(jp_name_str)
                                else:
                                    name_table_internal[jp_name_str] = str(cn_name)
                        else:
                            LOGGER.warning(
                                f"CSV name替换表 {path} 中发现格式不正确的行 (行号 {row_idx}): {row}"
                            )
                file_loaded_successfully_internal = True
            else:
                LOGGER.warning(
                    f"不支持的 name替换表 文件格式: {file_extension}. 请使用 .xlsx 或 .csv 文件。"
                )
                return {}, [], False

        except FileNotFoundError:
            LOGGER.warning(f"name替换表文件未找到: {path}")
            return {}, [], False
        except Exception as e:
            LOGGER.error(f"载入name替换表 '{path}' 时出错: {e}")
            return {}, [], False

        return (
            name_table_internal,
            missing_cn_names_internal,
            file_loaded_successfully_internal,
        )

    # First attempt to load
    name_table, missing_cn_names, file_loaded_successfully = _load_internal(
        name_table_path
    )

    table_base_name = os.path.basename(name_table_path)
    # Check for missing CN_Names after the first attempt
    if file_loaded_successfully and missing_cn_names and firstime_load:
        LOGGER.warning(
            f"\n\n(这个提示只会在首次显示)\n\n'{table_base_name}' 中有name的翻译未补齐，可以现在编辑并补齐对应翻译，或以后编辑并通过刷新结果来补全name字段的翻译。\n\n配置文件中usePostDictInName, useGPTDictInName也可将译后、GPT字典用于刷写name字段。"
        )
        print()
        try:
            input("按 Enter 继续，或ctrl+c暂时返回...")
        except EOFError:
            raise KeyboardInterrupt
        # Second attempt to load after user edit
        name_table, missing_cn_names, file_loaded_successfully = _load_internal(
            name_table_path
        )

    # Log final status
    if file_loaded_successfully:
        LOGGER.info(f"{table_base_name} 载入 {len(name_table)} 条name替换表")
        # Check again after reload
        if missing_cn_names:
            LOGGER.warning(
                f"'{table_base_name}' 中name翻译有缺失。程序将继续，但这些名字不会被替换。"
            )

    name_counter = {}
    for chunk in chunks:
        for tran in chunk.trans_list:
            if tran.speaker and isinstance(tran.speaker, str):
                if tran.speaker not in name_counter:
                    name_counter[tran.speaker] = 0
                name_counter[tran.speaker] += 1

    name_counter = dict(sorted(name_counter.items(), key=lambda item: item[1], reverse=True))

    if usePreDictInName:
        count = 0
        for name in name_counter.keys():
            if pre_dic.get_dst(name)!= "":
                name_table[name] = pre_dic.get_dst(name)
                count += 1
        LOGGER.info(f"usePreDictInName: 使用译前字典载入 {count} 条name替换表")

    if useGPTDictInName:
        count = 0
        for name in name_counter.keys():
            if gpt_dic.get_dst(name) != "":
                name_table[name] = gpt_dic.get_dst(name)
                count += 1
        LOGGER.info(f"useGPTDictInName: 使用GPT字典载入 {count} 条name替换表")

    if usePostDictInName:
        count = 0
        for name in name_counter.keys():
            if post_dic.get_dst(name) != "":
                name_table[name] = post_dic.get_dst(name)
                count += 1
        LOGGER.info(f"usePostDictInName: 使用译后字典载入 {count} 条name替换表")

    return name_table


async def dump_name_table_from_chunks(
    chunks: List[SplitChunkMetadata], proj_config: CProjectConfig
):
    name_dict = {}
    proj_dir = proj_config.getProjectDir()
    gpt_dic = proj_config.gpt_dic

    for chunk in chunks:
        for tran in chunk.trans_list:
            if tran.speaker and isinstance(tran.speaker, str):
                if tran.speaker not in name_dict:
                    name_dict[tran.speaker] = 0
                name_dict[tran.speaker] += 1

    name_dict = dict(sorted(name_dict.items(), key=lambda item: item[1], reverse=True))

    LOGGER.debug(f"共发现 {len(name_dict)} 个人名，按出现次数排序如下：")
    for name, count in name_dict.items():
        LOGGER.debug(f"{name}: {count}")

    # Ask user for export format
    try:
        export_format = await inquirer.select(
            message="请选择导出 name替换表 的格式 (这个替换表可以刷写结果文件中的name字段):",
            choices=[
                Choice(value="csv", name="CSV (默认)"),
                Choice(value="xlsx", name="Excel (.xlsx)"),
            ],
            default="csv",
        ).execute_async()
    except Exception as e:
        LOGGER.warning(f"无法获取用户输入，将默认使用 CSV 格式: {e}")
        export_format = "csv"  # Default to csv if inquirer fails

    file_extension = f".{export_format}"
    output_path = joinpath(proj_dir, f"name替换表{file_extension}")

    try:
        if export_format == "xlsx":
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "NameTable"

            # Write header
            sheet["A1"] = "JP_Name"
            sheet["B1"] = "CN_Name"
            sheet["C1"] = "Count"

            # Write data
            row_num = 2
            for name, count in name_dict.items():
                sheet[f"A{row_num}"] = name
                sheet[f"B{row_num}"] = ""
                sheet[f"C{row_num}"] = count
                row_num += 1

            workbook.save(output_path)
            LOGGER.info(
                f"name已保存到'{output_path}' (Excel格式)，填入CN_Name后可用于后续翻译name字段。"
            )
        elif export_format == "csv":
            with open(output_path, "w", newline="", encoding="utf-8-sig") as csvfile:
                writer = csv.writer(csvfile)
                # Write header
                writer.writerow(["JP_Name", "CN_Name", "Count"])
                # Write data
                for name, count in name_dict.items():
                    writer.writerow([name, "", count])
            LOGGER.info(
                f"name已保存到'{output_path}' (CSV格式)，填入CN_Name后可用于后续翻译name字段。"
            )

    except Exception as e:
        LOGGER.error(f"保存name替换表到 {export_format.upper()} 时出错: {e}")
